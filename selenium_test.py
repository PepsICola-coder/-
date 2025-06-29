#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time       : 2025/5/24 11:09
# @Author     : ZZ
# @File       : selenium_test.py
# @Software   : PyCharm
# !/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName: uc.py
# @Time: 2025/5/19 下午1:36
# @Author: ZZ
# type: ignore
from undetected_chromedriver import Chrome, ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random


# 新增：创建driver的独立函数
def create_driver(driver_path):
    options = ChromeOptions()
    # 增强反爬措施
    options.add_argument("--disable-blink-features=AutomationControlled")  # 禁用Blink控制特征
    options.add_argument("--disable-infobars")  # 禁用信息栏
    options.add_argument("--disable-notifications")  # 禁用通知
    options.add_argument("--disable-popup-blocking")  # 禁用弹窗拦截
    options.add_argument("--disable-save-password-bubble")  # 禁用密码保存提示
    options.add_argument("--disable-gpu")  # 禁用GPU加速
    # 窗口设置改为随机尺寸（原固定最大化设置）
    options.add_argument(f"--window-size={random.randint(1000, 1400)},{random.randint(800, 1000)}")

    # 添加驱动配置（禁用日志）
    service_args = [
        '--disable-logging',
        '--silent'
    ]
    driver = Chrome(
        options=options,
        driver_executable_path=driver_path,
        service_args=service_args
    )
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

# 修改后的主函数
def auto_visit_baidu(driver_path):
    # 创建driver实例
    driver = create_driver(driver_path)

    return driver# 调用独立创建函数
    
    # 封装网站访问操作到函数
def visit_website(driver_path, keywords_list):
    selected_driver = create_driver(driver_path)
    selected_driver.get("https://www.zhuangxiaozhuang.com")
    original_window = selected_driver.current_window_handle
    result_urls = []  # 新增URL存储列表

    try:
        # 新增：遍历关键词列表
        for keyword in keywords_list:
            try:
                # 新增：清空输入框（针对连续搜索场景）
                search_input = WebDriverWait(selected_driver, 10).until(
                    EC.presence_of_element_located((By.ID, "search-input"))
                )
                search_input.clear()
                
                # 修改：使用当前循环的关键词
                search_input.send_keys(keyword)
                # 点击前随机延迟
                time.sleep(random.uniform(0.2, 1.5))

                search_button = selected_driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div[2]/form/div/div[1]/button")
                search_button.click()

                # 处理页面跳转后的新内容
                WebDriverWait(selected_driver, 10).until(
                    EC.url_contains("https://www.zhuangxiaozhuang.com/index.php?s=/index/search")
                )
                # 跳转后随机延迟
                time.sleep(random.uniform(1.0, 4.0))

                # 新增：定位并点击搜索结果中的第一个商品（示例XPath）
                # 使用显式等待确保元素加载完成
                first_product = selected_driver.find_element(By.XPATH,'//*[@id="app"]/div/div/div/div/div[3]/div/div')
                first_product.click()

                # 新增：等待新窗口出现并切换
                WebDriverWait(selected_driver, 15).until(EC.number_of_windows_to_be(2))  # 等待新窗口出现
                new_window = [window for window in selected_driver.window_handles if window != original_window][0]
                selected_driver.switch_to.window(new_window)  # 切换到新窗口

                WebDriverWait(selected_driver, 15).until(  # 延长等待时间至15秒
                    EC.url_contains("https://www.zhuangxiaozhuang.com/index.php?s=/index/goods/index/id")  # 改为等待URL包含商品路径特征
                )

                # 新增：添加商品详情页主体元素检查（使用CSS选择器）
                WebDriverWait(selected_driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '#app > div > div > div > div > div.goods_images_base_best-box > div.goods_base-box'))  # 请替换实际CSS选择器
                )

                # 获取当前页面URL并存储到列表（替换原有打印逻辑）
                current_url = selected_driver.current_url
                result_urls.append(current_url)

                # 关闭当前窗口并切换回原始窗口（保持原有逻辑）
                selected_driver.close()
                selected_driver.switch_to.window(original_window)
                
            except Exception as e:
                print(f"处理关键词[{keyword}]时出错: {str(e)}")
                result_urls.append("ERROR")
                continue  # 继续处理下一个关键词

    finally:
        # 修改：在所有关键词处理完成后退出
        selected_driver.quit()
    
    return result_urls  # 新增返回结果列表

# 新增Excel读取函数（需要安装pandas和openpyxl）
def read_excel_column(file_path, column_index=0):
    """
    读取xlsx文件指定列数据并返回列表
    参数：
        file_path: Excel文件路径
        column_index: 列索引（从0开始）
    返回：
        list: 包含该列所有数据的列表
    """
    try:
        import pandas as pd
    except ImportError:
        raise RuntimeError("需要安装pandas库，请执行: pip install pandas openpyxl")

    try:
        # 读取Excel文件（使用openpyxl引擎）
        df = pd.read_excel(file_path, engine='openpyxl')
        
        # 验证列索引有效性
        if column_index < 0 or column_index >= len(df.columns):
            raise ValueError(f"无效列索引，可用列范围：0-{len(df.columns)-1}")
            
        # 转换为列表并去除空值
        column_data = df.iloc[:, column_index].dropna().tolist()
        return column_data
        
    except FileNotFoundError:
        raise FileNotFoundError(f"文件不存在：{file_path}")
    except Exception as e:
        raise RuntimeError(f"读取Excel失败：{str(e)}")

# 在read_excel_column函数后新增写入列的方法
def write_excel_column(file_path, data_list, target_column_index=0, sheet_name='Sheet1'):
    """
    将数据列表写入xlsx文件指定列
    """
    try:
        from openpyxl import load_workbook
        import pandas as pd
    except ImportError:
        raise RuntimeError("需要安装pandas和openpyxl库，请执行: pip install pandas openpyxl")

    try:
        # 使用openpyxl加载现有工作簿保留格式
        book = load_workbook(file_path)
        # 获取或创建工作表
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)

        # 修复行数获取方式：使用max_row替代len(sheet.rows)
        max_write_rows = sheet.max_row - 1 if sheet.max_row >= 1 else 0
        
        # 从第二行开始写入（保持标题行不变）
        adjusted_data = data_list[:max_write_rows]
        
        # 写入数据到指定列（保留原有单元格格式）
        for i, value in enumerate(adjusted_data, start=2):  # 从第2行开始
            cell = sheet.cell(row=i, column=target_column_index+1)
            cell.value = value
            
        # 保存时保留原文件格式
        book.save(file_path)
        return True

    except PermissionError:
        raise RuntimeError("文件被其他程序占用，请关闭后重试")
    except Exception as e:
        raise RuntimeError(f"写入Excel失败：{str(e)}")

if __name__ == "__main__":
    # 示例用法（读取第1列数据）
    # 注意：实际使用时请取消注释
    data = read_excel_column(r"F:\兼职\福满堂\福满堂上架表6.11.xlsx", column_index=3)
    print("读取到的列数据:", data)
    input("......")
    # 原有调用保持不变
    ls = visit_website(r"F:\ChromeDownload\chromedriver-win64\chromedriver.exe",data)
    write_excel_column(r"F:\兼职\福满堂\福满堂上架表6.11.xlsx", ls, target_column_index=36, sheet_name='Sheet1')